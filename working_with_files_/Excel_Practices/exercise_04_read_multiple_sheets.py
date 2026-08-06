from openpyxl import load_workbook

workbook = load_workbook("bim_multi_sheet_report.xlsx")

print("Available worksheets:")
print(workbook.sheetnames)

doors = workbook["Doors"]
windows = workbook["Windows"]
rooms = workbook["Rooms"]

fire_rating = doors["D2"].value
print(f"Door Fire Rating: {fire_rating}")

window_level = windows["C2"].value
print(f"Window Level: {window_level}")

room_name = rooms["B2"].value
print(f"Room Name: {room_name}")

print("Rooms data:")

for row in rooms.iter_rows(min_row=2, values_only=True):
    print(row)
