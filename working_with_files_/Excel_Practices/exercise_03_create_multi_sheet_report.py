from openpyxl import Workbook

workbook = Workbook()

doors = workbook.active
doors.title = "Doors"

windows = workbook.create_sheet("Windows")
rooms = workbook.create_sheet("Rooms")

doors["A1"] = "Element ID"
doors["B1"] = "Family Name"
doors["C1"] = "Level"
doors["D1"] = "Fire Rating"

doors["A2"] = 3001
doors["B2"] = "Single Flush 900mm"
doors["C2"] = "Level 1"
doors["D2"] = "60min"

windows["A1"] = "Element ID"
windows["B1"] = "Family Name"
windows["C1"] = "Level"
windows["D1"] = "Glazing Type"

windows["A2"] = 4001
windows["B2"] = "Fixed Window 1200x1500"
windows["C2"] = "Level 2"
windows["D2"] = "Double Glazed"

rooms["A1"] = "Room Number"
rooms["B1"] = "Room Name"
rooms["C1"] = "Level"
rooms["D1"] = "Area"

rooms["A2"] = 101
rooms["B2"] = "Office"
rooms["C2"] = "Level 1"
rooms["D2"] = 24.5

workbook.save("bim_multi_sheet_report.xlsx")
print("Multi-sheet BIM Excel report created successfully.")
