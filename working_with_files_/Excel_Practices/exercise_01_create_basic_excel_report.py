from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.title = "Revit Elements"

ws["A1"] = "Element ID"
ws["B1"] = "Category"
ws["C1"] = "Type Name"
ws["D1"] = "Level"

ws["A2"] = 2001
ws["B2"] = "Door"
ws["C2"] = "Single Flush 900mm"
ws["D2"] = "Level 1"

wb.save("revit_elements_basic_report.xlsx")

print("Basic Excel report created successfully.")
