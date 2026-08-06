from openpyxl import load_workbook

workbook = load_workbook("revit_elements_basic_report.xlsx")
worksheet = workbook.active

element_id = worksheet["A2"].value
category = worksheet["B2"].value
type_name = worksheet["C2"].value
level = worksheet["D2"].value

print("Element ID:", element_id)
print("Category:", category)
print("Type Name:", type_name)
print("Level:", level)

print("Reading all data rows:")

for row in worksheet.iter_rows(min_row=2, values_only=True):
    print(row)
