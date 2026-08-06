from openpyxl import load_workbook

workbook = load_workbook("bim_multi_sheet_report.xlsx")

print("Worksheet Summary")
print("-----------------")

for worksheet in workbook.worksheets:
    print(f"Worksheet name: {worksheet.title}")
    print(f"Number of columns: {worksheet.max_column}")
    print(f"Number of rows: {worksheet.max_row}")

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=2,
        values_only=True
    ):
        print(f"First data row: {row}")

    print()
